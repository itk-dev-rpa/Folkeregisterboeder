"""This module is responsible for reading and writing Excel files in the correct format."""

from io import BytesIO

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from robot_framework.sub_process.task import Task
from robot_framework.sub_process import sap


def read_excel(file: BytesIO, case_worker_name: str) -> tuple[Task, ...]:
    """Read an Excel sheet and output a list of Task objects.

    The columns of the Excel sheet is expected to be:
    Dato, Flyttedato, BLANK , Anmeldelsesdato, Sagsnr., Flyttetype, Status, CPR-nr., Navn, Brev, Faktura, PDF, Bemærk


    Args:
        file: The Excel file as an BytesIO object.

    Returns:
        A tuple of Task objects with length equal to the number of rows in Excel.
    """

    input_sheet: Worksheet = load_workbook(file, read_only=True).active

    task_list = []

    iter_ = iter(input_sheet)
    next(iter_)  # Skip header row
    for row in iter_:
        task = Task(
            case_worker_name=case_worker_name,
            task_date=row[0].value,
            move_date=row[1].value,
            register_date=row[3].value,
            eflyt_case_number=str(row[4].value),
            eflyt_categories=row[5].value,
            eflyt_status=row[6].value,
            cpr=str(row[7].value),
            name=row[8].value
        )
        task_list.append(task)

    return tuple(task_list)


def write_excel(tasks: list[Task]) -> BytesIO:
    """Write a list of task objects to an excel sheet.

    Args:
        tasks: The list of task objects to write.

    Returns:
        A BytesIO object containing the Excel sheet.
    """
    wb = Workbook()
    sheet: Worksheet = wb.active

    header = ["Dato", "Flyttedato", "", "Anmeldelsesdato", "Sagsnr.", "Flyttetype", "Status", "CPR-nr.", "Navn", "Brev dato", "Opkrævning dato", "Journalisering dato", "Bøde beløb"]
    sheet.append(header)

    for task in tasks:
        fine_rate = sap.get_fine_rate(task.move_date) if task.invoice_date else None
        row = [task.task_date, task.move_date, "", task.register_date, task.eflyt_case_number, task.eflyt_categories, task.eflyt_status, task.cpr, task.name, task.letter_date, task.invoice_date, task.journal_date, fine_rate]
        sheet.append(row)

    file = BytesIO()
    wb.save(file)
    return file
